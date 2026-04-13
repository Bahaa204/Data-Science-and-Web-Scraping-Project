# Importing Libraries
from pathlib import Path
import pandas as pd
import numpy as np

# Importing Types and Functions
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, Index
import Constants
from Helpers import CLEAN_DATA, ANALYZED_DATA


def main() -> None:

    subs: DataFrame = pd.read_csv(CLEAN_DATA / "subscriptions_cleaned.csv")
    hourly: DataFrame = pd.read_csv(CLEAN_DATA / "hourly_wage_cleaned.csv")
    steam: DataFrame = pd.read_csv(CLEAN_DATA / "steam_games_cleaned.csv")

    subs.columns = subs.columns.str.strip()
    hourly.columns = hourly.columns.str.strip()
    steam.columns = steam.columns.str.strip()

    steam["Region Name"] = (
        steam["Region"].map(Constants.region_names).fillna(steam["Region"])
    )

    subs["plan_tier"] = (
        subs["plan_type"].map(Constants.tier_map).fillna(subs["plan_type"])
    )

    # Analysis 1 — Streaming Services Affordability

    affordability: DataFrame = (
        subs[subs["price"] > 0]
        .merge(
            hourly[["country_name", "hourly_wage_usd"]],
            left_on="region",
            right_on="country_name",
            how="left",
        )
        .dropna(subset=["hourly_wage_usd"])
        .copy()
    )

    affordability["hours_to_afford"] = (
        affordability["price"] / affordability["hourly_wage_usd"]
    ).round(2)

    affordability["affordability_tier"] = pd.cut(
        affordability["hours_to_afford"],
        bins=[0, 1, 5, 15, 50, 9999],
        labels=[
            "Very Affordable (<1h)",
            "Affordable (1-5h)",
            "Moderate (5-15h)",
            "Expensive (15-50h)",
            "Very Expensive (>50h)",
        ],
    ).astype(str)

    aff_out: DataFrame = affordability[
        [
            "region",
            "platform",
            "plan_type",
            "price",
            "hourly_wage_usd",
            "hours_to_afford",
            "affordability_tier",
        ]
    ].copy()

    aff_out.columns = [
        "Region",
        "Platform",
        "Plan Type",
        "Price (USD)",
        "Hourly Wage (USD)",
        "Hours to Afford",
        "Affordability Tier",
    ]
    aff_out = aff_out.sort_values("Hours to Afford")

    aff_pivot: DataFrame = (
        affordability.pivot_table(
            index="region",
            columns="platform",
            values="hours_to_afford",
            aggfunc="median",
        )
        .round(2)
        .reset_index()
    )
    aff_pivot.columns.name = None

    # Analysis 2 — Steam Pricing

    steam_region: DataFrame = (
        steam.groupby("Region Name")
        .agg(
            Avg_Price=("Current Price", "mean"),
            Avg_Original=("Original Price", "mean"),
            Avg_Discount=("Discount", "mean"),
            Game_Count=("Title", "count"),
        )
        .round(2)
        .reset_index()
    )
    steam_region.columns = [
        "Region",
        "Avg Current Price (USD)",
        "Avg Original Price (USD)",
        "Avg Discount (%)",
        "Game Count",
    ]

    steam_top_deals: DataFrame = steam[steam["Discount"] < 0].copy()
    steam_top_deals["Discount Abs"] = steam_top_deals["Discount"].abs()
    steam_top_deals = (
        steam_top_deals.sort_values("Discount Abs", ascending=False)
        .drop_duplicates("Title")
        .head(50)[
            [
                "Title",
                "Region Name",
                "Current Price",
                "Original Price",
                "Discount",
                "Platforms",
            ]
        ]
        .copy()
    )
    steam_top_deals.columns = [
        "Title",
        "Region",
        "Current Price (USD)",
        "Original Price (USD)",
        "Discount (%)",
        "Platforms",
    ]

    steam_pivot_price: DataFrame = (
        steam.pivot_table(
            index="Title",
            columns="Region Name",
            values="Current Price",
            aggfunc="first",
        )
        .round(2)
        .reset_index()
    )
    steam_pivot_price.columns.name = None
    num_c: Index = steam_pivot_price.select_dtypes(include="number").columns
    steam_pivot_price["Price Range (USD)"] = (
        steam_pivot_price[num_c].max(axis=1) - steam_pivot_price[num_c].min(axis=1)
    ).round(2)
    steam_pivot_price = steam_pivot_price.sort_values(
        "Price Range (USD)", ascending=False
    ).head(100)

    # Analysis 3 — Subscription Comparison

    plat_tier: DataFrame = (
        subs[subs["price"] > 0]
        .pivot_table(
            index="platform", columns="plan_tier", values="price", aggfunc="median"
        )
        .round(2)
        .reset_index()
    )
    plat_tier.columns.name = None

    reg_plat: DataFrame = (
        subs[subs["price"] > 0]
        .pivot_table(
            index="region", columns="platform", values="price", aggfunc="median"
        )
        .round(2)
        .reset_index()
    )
    reg_plat.columns.name = None
    num_cols: Index = reg_plat.select_dtypes(include="number").columns
    reg_plat["Cheapest Platform"] = reg_plat[num_cols].idxmin(axis=1)
    reg_plat["Most Expensive Platform"] = reg_plat[num_cols].idxmax(axis=1)

    plat_disp: DataFrame = (
        subs[subs["price"] > 0]
        .groupby(["platform", "plan_tier"])
        .agg(
            Min_Price=("price", "min"),
            Max_Price=("price", "max"),
            Median_Price=("price", "median"),
            Regions=("region", "nunique"),
        )
        .round(2)
        .reset_index()
    )
    plat_disp.columns = [
        "Platform",
        "Plan Tier",
        "Min Price (USD)",
        "Max Price (USD)",
        "Median Price (USD)",
        "Regions Available",
    ]
    plat_disp["Price Ratio (Max/Min)"] = (
        plat_disp["Max Price (USD)"] / plat_disp["Min Price (USD)"].replace(0, np.nan)
    ).round(2)

    # Analysis 4 — Cross-dataset Affordability Summary

    cross: DataFrame = (
        affordability.groupby("region")
        .agg(
            Avg_Sub_Price=("price", "mean"),
            Hourly_Wage=("hourly_wage_usd", "first"),
            Avg_Hours=("hours_to_afford", "mean"),
            Best_Platform=(
                "platform",
                lambda x: affordability.loc[x.index]
                .groupby("platform")["hours_to_afford"]
                .mean()
                .idxmin(),
            ),
            Min_Hours=("hours_to_afford", "min"),
        )
        .round(2)
        .reset_index()
    )
    cross.columns = [
        "Region",
        "Avg Sub Price (USD)",
        "Hourly Wage (USD)",
        "Avg Hours to Afford",
        "Most Affordable Platform",
        "Min Hours to Afford",
    ]
    cross = cross.sort_values("Avg Hours to Afford")

    # Style Helpers

    def fill(h: Constants.Hue) -> PatternFill:
        return PatternFill("solid", start_color=h, fgColor=h)

    def border() -> Border:
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def center() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left() -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=False)

    def header(worksheet, r, cols, bg: Constants.NAV) -> None:
        for ci, v in enumerate(cols, 1):
            c = worksheet.cell(row=r, column=ci, value=v)
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = border()

    color_alt: Constants.ALT = "EBF3FB"
    color_white: Constants.WHITE = "FFFFFF"

    def row_style(worksheet: Worksheet, r, n, alt: bool = False) -> None:
        for ci in range(1, n + 1):
            c: Cell | MergedCell = worksheet.cell(row=r, column=ci)
            c.fill = fill(color_alt if alt else color_white)
            c.border = border()
            c.font = Font(name="Arial", size=10)

    def autofit(worksheet: Worksheet, mn: int = 10, mx: int = 36) -> None:
        for col in worksheet.columns:
            best: int = mn
            for cell in col:
                try:
                    v: str = str(cell.value) if cell.value is not None else ""
                    best = max(best, min(len(v) + 2, mx))
                except:
                    pass
            if col[0].column:
                worksheet.column_dimensions[get_column_letter(col[0].column)].width = (
                    best
                )

    def write_df(
        worksheet: Worksheet, df: DataFrame, hdr_bg: Constants.NAV, start_row: int = 2
    ) -> None:
        header(worksheet, start_row, df.columns.tolist(), bg=hdr_bg)
        for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
            row_style(worksheet, ri, len(df.columns), alt=(ri % 2 == 0))
            for ci, val in enumerate(row, 1):
                c: Cell | MergedCell = worksheet.cell(row=ri, column=ci)
                if isinstance(c, MergedCell):
                    continue
                if isinstance(val, (float, np.floating)) and not np.isnan(float(val)):
                    c.value = float(val)
                    c.number_format = "#,##0.00"
                elif isinstance(val, (int, np.integer)):
                    c.value = int(val)
                else:
                    c.value = str(val) if pd.notna(val) else ""
                c.alignment = (
                    center()
                    if isinstance(val, (float, int, np.floating, np.integer))
                    else left()
                )

    def title_row(
        worksheet: Worksheet, title: str, ncols: int, bg: Constants.BLUE
    ) -> None:
        worksheet.row_dimensions[1].height = 24
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c: Cell | MergedCell = worksheet.cell(row=1, column=1, value=title)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
        c.fill = fill(bg)
        c.alignment = center()

    def color_scale(
        worksheet: Worksheet, col_letter: str, start_row: int, end_row: int
    ) -> None:
        worksheet.conditional_formatting.add(
            f"{col_letter}{start_row}:{col_letter}{end_row}",
            ColorScaleRule(
                start_type="min",
                start_color="70AD47",
                mid_type="percentile",
                mid_value=50,
                mid_color="F4B942",
                end_type="max",
                end_color="FF6B6B",
            ),
        )

    # build workbook

    workbook: Workbook = Workbook()

    # Sheet 1 — Combined Raw Data

    worksheet1: Worksheet | None = workbook.active

    if not worksheet1:
        raise Exception("Worksheet Not Found")

    worksheet1.title = "Combined Data"

    steam_f: DataFrame = steam[
        [
            "Region Name",
            "Title",
            "Platforms",
            "Current Price",
            "Original Price",
            "Discount",
            "Scraped At",
            "currency",
        ]
    ].copy()

    steam_f.columns = [
        "steam_region",
        "steam_name",
        "steam_platforms",
        "steam_current_price_usd",
        "steam_original_price_usd",
        "steam_discount_pct",
        "steam_scraped_at",
        "steam_currency",
    ]

    subs_f: DataFrame = subs[
        ["platform", "region", "plan_type", "price", "scraped_at", "currency"]
    ].copy()

    subs_f.columns = [
        "sub_platform",
        "sub_region",
        "sub_plan_type",
        "sub_price_usd",
        "sub_scraped_at",
        "sub_currency",
    ]

    max_rows: int = max(len(steam_f), len(subs_f))

    combined: DataFrame = pd.concat(
        [
            steam_f.reset_index(drop=True).reindex(range(max_rows)),
            subs_f.reset_index(drop=True).reindex(range(max_rows)),
        ],
        axis=1,
    )

    title_row(
        worksheet1,
        "Combined Raw Data — Steam Games & Subscriptions",
        len(combined.columns),
        "2E75B6",
    )

    write_df(worksheet1, combined, "1F3864")

    worksheet1.freeze_panes = "A3"

    autofit(worksheet1)

    # Sheet 2 — Analysis 1: Affordability Detail

    worksheet2: Worksheet = workbook.create_sheet("A1 - Affordability Detail")

    title_row(
        worksheet2,
        "Analysis 1 - Hours of Work to Afford Subscriptions (Detail)",
        len(aff_out.columns),
        "2E75B6",
    )

    write_df(worksheet2, aff_out, "1F3864")

    worksheet2.freeze_panes = "A3"

    autofit(worksheet2)
    color_scale(worksheet2, "F", 3, len(aff_out) + 2)

    # Sheet 3 — Analysis 1: Affordability Pivot
    worksheet3 = workbook.create_sheet("A1 - Affordability Pivot")

    title_row(
        worksheet3,
        "Analysis 1 - Median Hours to Afford by Region & Platform",
        len(aff_pivot.columns),
        "2E75B6",
    )

    write_df(worksheet3, aff_pivot, "1F3864")

    worksheet3.freeze_panes = "B3"

    autofit(worksheet3)

    for i in range(2, len(aff_pivot.columns) + 1):
        color_scale(worksheet3, get_column_letter(i), 3, len(aff_pivot) + 2)

    # Sheet 4 — Analysis 2: Steam Regional Pricing

    worksheet4: Worksheet = workbook.create_sheet("A2 - Steam Regional Pricing")

    title_row(
        worksheet4,
        "Analysis 2 - Steam Games: Regional Pricing Summary",
        len(steam_region.columns),
        "2E75B6",
    )

    write_df(worksheet4, steam_region, "1F3864")

    worksheet4.freeze_panes = "A3"

    autofit(worksheet4)

    # Sheet 5 — Analysis 2: Top Steam Deals
    worksheet5: Worksheet = workbook.create_sheet("A2 - Top Steam Deals")

    title_row(
        worksheet5,
        "Analysis 2 - Top 50 Steam Games by Discount %",
        len(steam_top_deals.columns),
        "2E75B6",
    )
    write_df(worksheet5, steam_top_deals, "1F3864")

    worksheet5.freeze_panes = "A3"
    autofit(worksheet5)

    color_scale(worksheet5, "E", 3, len(steam_top_deals) + 2)

    # Sheet 6 — Analysis 2: Steam Cross-Region Prices
    worksheet6: Worksheet = workbook.create_sheet("A2 - Steam Price by Region")
    title_row(
        worksheet6,
        "Analysis 2 - Game Prices Across Regions (Top 100 by Price Range)",
        len(steam_pivot_price.columns),
        "2E75B6",
    )
    write_df(worksheet6, steam_pivot_price, "1F3864")
    worksheet6.freeze_panes = "B3"
    autofit(worksheet6)

    # Sheet 7 — Analysis 3: Platform vs Plan Tier
    worksheet7: Worksheet = workbook.create_sheet("A3 - Platform vs Plan Tier")
    title_row(
        worksheet7,
        "Analysis 3 - Median Price (USD) by Platform & Plan Tier",
        len(plat_tier.columns),
        "2E75B6",
    )
    write_df(worksheet7, plat_tier, "1F3864")
    worksheet7.freeze_panes = "B3"
    autofit(worksheet7)

    # Sheet 8 — Analysis 3: Region vs Platform
    worksheet8: Worksheet = workbook.create_sheet("A3 - Region vs Platform")
    title_row(
        worksheet8,
        "Analysis 3 - Median Subscription Price (USD) by Region & Platform",
        len(reg_plat.columns),
        "2E75B6",
    )
    write_df(worksheet8, reg_plat, "1F3864")
    worksheet8.freeze_panes = "B3"
    autofit(worksheet8)

    # Sheet 9 — Analysis 3: Price Dispersion
    worksheet9: Worksheet = workbook.create_sheet("A3 - Price Dispersion")
    title_row(
        worksheet9,
        "Analysis 3 - Subscription Price Dispersion by Platform & Plan Tier",
        len(plat_disp.columns),
        "2E75B6",
    )
    write_df(worksheet9, plat_disp, "1F3864")
    worksheet9.freeze_panes = "A3"
    autofit(worksheet9)

    # Sheet 10 — Analysis 4: Cross-Dataset Affordability Summary
    worksheet10: Worksheet = workbook.create_sheet("A4 - Affordability Summary")
    title_row(
        worksheet10,
        "Analysis 4 - Cross-Dataset Affordability Summary by Region",
        len(cross.columns),
        "2E75B6",
    )
    write_df(worksheet10, cross, "1F3864")
    worksheet10.freeze_panes = "A3"
    autofit(worksheet10)
    color_scale(worksheet10, "D", 3, len(cross) + 2)

    OUTPUT_PATH: Path = ANALYZED_DATA / "all_clean_data_combined_analyzed.xlsx"
    workbook.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(
        f"Sheets ({len(workbook.worksheets)}): {[s.title for s in workbook.worksheets]}"
    )


if __name__ == "__main__":
    main()
